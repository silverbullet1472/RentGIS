from openpyxl import load_workbook
import psycopg2
from geo_coder import *

if __name__ == "__main__":
    conn = psycopg2.connect(dbname="rent_db", user="postgres", password="postgresql", host="127.0.0.1", port="5432")
    cur = conn.cursor()
    # 1 建立新表
    cur.execute("""
                  create table poi_processed
                  (
                      leixing          varchar(50),
                      mingcheng        varchar(50),
                      dizhi            varchar(50),
                      quxian           varchar(50),
                      yilei            varchar(50),
                      erlei            varchar(50),
                      
                      bd_lon           varchar(20),
                      bd_lat           varchar(20),
                      wgs_lon          varchar(20),
                      wgs_lat          varchar(20)
                  );
                  """)
    conn.commit()

    table = []
    wb = load_workbook("宜昌市.xlsx", read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    for name in name_list:
        my_sheet = wb[name]
        for row in my_sheet.iter_rows(min_row=2):
            entry = []
            for cell in row:
                entry.append(str(cell.value)[0:50])
            print(entry)
            location = bd09_to_wgs84(float(entry[2]), float(entry[1]))
            reodered_row = [str(name),entry[0], entry[3], entry[4], entry[5], entry[6], entry[2], entry[1], location[0], location[1]]
            print(reodered_row)
            table.append(reodered_row)

    cur.executemany(f"insert into poi_processed values (%s,%s,%s,%s,%s, %s,%s,%s,%s,%s)", table)
    conn.commit()
    conn.close()